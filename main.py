import csv
import pathlib
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
import doctest
import cProfile
profile = cProfile.Profile()


class InCon:
	"""Собирает информации из консоли и связывает класса DataSet и Report
    Attributes:
           file_name (str): Название файла
           vacancy_name (str): Название вакансии"""

	def __init__(self):
		"""Инициализирует класс InputConnect, берёт информацию из консоли, создаёт экземпляр класса,
        собирает статистику, выводит её и позволяет сгенерировать excel-таблицу
        """

		self.file_name = input('Введите название файла: ')
		self.vacancy_name = input('Введите название профессии: ')
		profile.enable()
		dataset = DataSet(self.file_name, self.vacancy_name)
		stats1, stats2, stats3, stats4, stats5, stats6 = dataset.get_statistic()
		dataset.print_statistic(stats1, stats2, stats3, stats4, stats5, stats6)

		report = Report(self.vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6)
		report.generate_image()
		# report.generate_pdf()


class Report:
	"""Генерирует excel-файл
    Attributes:
       wb (Workbook): Экземпляр класса Workbook
       vacancy_name (str): Название вакансии
       stats1: Динамика уровня зарплат по годам
       stats2: Динамика количества вакансий по годам
       stats3: Динамика уровня зарплат по годам для профессии
       stats4: Динамика количества вакансий по годам для профессии
       stats5: Уровень зарплат по городам (в порядке убывания)
       stats6: Доля вакансий по городам (в порядке убывания)"""

	def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
		"""Инициализирует класс Report и создаёт экземпляр класс Workbook, который создает exel таблицу

		>>> Report('Аналитик', {}, {}, {}, {}, {}, {}).vacancy_name
		'Аналитик'
		>>> type(Report('', {}, {}, {}, {}, {}, {})).__name__
		'Report'
		>>> type(Report('', {}, {}, {}, {}, {}, {}).wb).__name__
		'Workbook'
		>>> type(Report('', {}, {}, {}, {}, {}, {}).stats1).__name__
		'dict'
		>>> Report('', {2007: 30000, 2008: 35000, 2009: 40000}, {}, {}, {}, {}, {}).stats1[2007]
		30000
		>>> Report('', {}, {}, {}, {}, {}, {'Москва': 0.441, 'Санкт-Петербург': 0.223}).stats6['Москва']
		0.441
		"""
		self.wb = Workbook()
		self.vacancy_name = vacancy_name
		self.stats1 = stats1
		self.stats2 = stats2
		self.stats3 = stats3
		self.stats4 = stats4
		self.stats5 = stats5
		self.stats6 = stats6

	def generate_excel(self):
		"""Генерирует файл report.xlsx, где выводится вся собранна статистика,
         а также сохраняет файл report.xlsx в директорию,
         откуда запускается Python-скрипт"""

		ws1 = self.wb.active
		ws1.title = 'Статистика по годам'
		ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий',
		            'Количество вакансий - ' + self.vacancy_name])
		for year in self.stats1.keys():
			ws1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

		data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий',
		         ' Количество вакансий - ' + self.vacancy_name]]
		column_widths = []
		for row in data:
			for i, cell in enumerate(row):
				if len(column_widths) > i:
					if len(cell) > column_widths[i]:
						column_widths[i] = len(cell)
				else:
					column_widths += [len(cell)]

		for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
			ws1.column_dimensions[get_column_letter(i)].width = column_width + 2

		data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
		for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
			data.append([city1, value1, '', city2, value2])
		ws2 = self.wb.create_sheet('Статистика по городам')
		for row in data:
			ws2.append(row)

		column_widths = []
		for row in data:
			for i, cell in enumerate(row):
				cell = str(cell)
				if len(column_widths) > i:
					if len(cell) > column_widths[i]:
						column_widths[i] = len(cell)
				else:
					column_widths += [len(cell)]

		for i, column_width in enumerate(column_widths, 1):
			ws2.column_dimensions[get_column_letter(i)].width = column_width + 2

		font_bold = Font(bold=True)
		for col in 'ABCDE':
			ws1[col + '1'].font = font_bold
			ws2[col + '1'].font = font_bold

		for index, _ in enumerate(self.stats5):
			ws2['E' + str(index + 2)].number_format = '0.00%'

		thin = Side(border_style='thin', color='00000000')

		for row in range(len(data)):
			for col in 'ABDE':
				ws2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

		for row, _ in enumerate(self.stats1):
			for col in 'ABCDE':
				ws1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

		self.wb.save(filename='report.xlsx')

	def generate_image(self):
		fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

		bar1 = ax1.bar(np.array(list(self.stats1.keys())) - 0.4, self.stats1.values(), width=0.4)
		bar2 = ax1.bar(np.array(list(self.stats1.keys())), self.stats3.values(), width=0.4)
		ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
		ax1.grid(axis='y')
		ax1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
		ax1.set_xticks(np.array(list(self.stats1.keys())) - 0.2, list(self.stats1.keys()), rotation=90)
		ax1.xaxis.set_tick_params(labelsize=8)
		ax1.yaxis.set_tick_params(labelsize=8)

		ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
		bar1 = ax2.bar(np.array(list(self.stats2.keys())) - 0.4, self.stats2.values(), width=0.4)
		bar2 = ax2.bar(np.array(list(self.stats2.keys())), self.stats4.values(), width=0.4)
		ax2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
		           prop={'size': 8})
		ax2.set_xticks(np.array(list(self.stats2.keys())) - 0.2, list(self.stats2.keys()), rotation=90)
		ax2.grid(axis='y')
		ax2.xaxis.set_tick_params(labelsize=8)
		ax2.yaxis.set_tick_params(labelsize=8)

		ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
		ax3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.stats5.keys()))]),
		         list(reversed(list(self.stats5.values()))), color='blue', height=0.5, align='center')
		ax3.yaxis.set_tick_params(labelsize=6)
		ax3.xaxis.set_tick_params(labelsize=8)
		ax3.grid(axis='x')

		ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
		other = 1 - sum([value for value in self.stats6.values()])
		ax4.pie(list(self.stats6.values()) + [other], labels=list(self.stats6.keys()) + ['Другие'],
		        textprops={'fontsize': 6})

		plt.tight_layout()
		plt.savefig('graph.png')

	def generate_pdf(self):
		env = Environment(loader=FileSystemLoader('.'))
		template = env.get_template("pdf_template.html")
		stats = []
		for year in self.stats1.keys():
			stats.append([year, self.stats1[year], self.stats2[year], self.stats3[year], self.stats4[year]])

		for key in self.stats6:
			self.stats6[key] = round(self.stats6[key] * 100, 2)

		pdf_template = template.render(
			{'name': self.vacancy_name, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
			 'stats': stats, 'stats5': self.stats5, 'stats6': self.stats6})
		config = pdfkit.configuration(wkhtmltopdf=r'D:\Games\wkhtmltopdf\bin\wkhtmltopdf.exe')
		pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})


class Vacancy:
	"""Класс для представления вакансии
    Attributes:
            name (str): Название вакансии
            salary_from (int): Нижняя граница вилки оклада
            salary_to (int): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
            area_name (str): Территория  вакансий
            salary_average (int): Среднее значение зарплаты
            year (int): Год публикации"""

	currency_to_rub = {
		"AZN": 35.68,
		"BYR": 23.91,
		"EUR": 59.90,
		"GEL": 21.74,
		"KGS": 0.76,
		"KZT": 0.13,
		"RUR": 1,
		"UAH": 1.64,
		"USD": 60.66,
		"UZS": 0.0055,
	}

	def __init__(self, vacancy):
		"""Инициализирует объект Vacancy, высчитывает среднюю зарплату и
         производит конвертацию для целочисленных полей
        Args:
            vacancy (dict): Словарь вакансии, из которого инициализируются переменные объекта
        >>> type(Vacancy({'name': '', 'salary_from': '0.1', 'salary_to': '0.1', 'salary_currency': 'RUR', 'area_name': '', 'published_at': '2007-12-03T17:34:36+0300'})).__name__
        'Vacancy'
        >>> Vacancy({'name': 'Аналитик', 'salary_from': '0.1', 'salary_to': '0.1', 'salary_currency': 'RUR', 'area_name': '', 'published_at': '2007-12-03T17:34:36+0300'}).name
        'Аналитик'
        >>> Vacancy({'name': '', 'salary_from': '0.1', 'salary_to': '0.1', 'salary_currency': 'RUR', 'area_name': '', 'published_at': '2007-12-03T17:34:36+0300'}).salary_from
        0
        >>> Vacancy({'name': '', 'salary_from': '0.1', 'salary_to': '0.1', 'salary_currency': 'RUR', 'area_name': '', 'published_at': '2007-12-03T17:34:36+0300'}).salary_to
        0
        >>> Vacancy({'name': '', 'salary_from': '0.1', 'salary_to': '0.1', 'salary_currency': 'RUR', 'area_name': '', 'published_at': '2007-12-03T17:34:36+0300'}).salary_currency
        'RUR'
        >>> Vacancy({'name': '', 'salary_from': '10', 'salary_to': '20', 'salary_currency': 'RUR', 'area_name': '', 'published_at': '2007-12-03T17:34:36+0300'}).salary_average
        15.0
        >>> Vacancy({'name': '', 'salary_from': '0.1', 'salary_to': '0.1', 'salary_currency': 'RUR', 'area_name': 'Москва', 'published_at': '2007-12-03T17:34:36+0300'}).area_name
        'Москва'
        >>> Vacancy({'name': '', 'salary_from': '0.1', 'salary_to': '0.1', 'salary_currency': 'RUR', 'area_name': '', 'published_at': '2007-12-03T17:34:36+0300'}).year
        2007
        """
		self.name = vacancy['name']
		self.salary_from = int(float(vacancy['salary_from']))
		self.salary_to = int(float(vacancy['salary_to']))
		self.salary_currency = vacancy['salary_currency']
		self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
		self.area_name = vacancy['area_name']
		self.year = int(vacancy['published_at'][:4])


class DataSet:
	"""Класс DataSet. В нем находятся методы по обработке данных,
     сбору статистику по вакансии и вывододу результата в консоль
    Attributes:
           file_name (str): Название файла
           vacancy_name (str): Название вакансии"""

	def __init__(self, file_name, vacancy_name):
		"""Инициализирует объект DataSet
        Args:
            file_name (str): Название файла
            vacancy_name (str): Название вакансии
        >>> type(DataSet('', '')).__name__
        'DataSet'
        >>> DataSet('vacancies.csv', 'Аналитик').file_name
        'vacancies.csv'
        >>> DataSet('vacancies.csv', 'Аналитик').vacancy_name
        'Аналитик'
        >>> type(DataSet('vacancies_csv', 'Аналитик').file_name).__name__
        'str'
        """

		self.file_name = file_name
		self.vacancy_name = vacancy_name

	@staticmethod
	def increment(dictionary, key, amount):
		"""Статический метод, который добавляет или же создаёт новый ключ в словаре
        Args:
            dictionary (dict): Словарь
            key (string or int): Ключ для этого словаря
            amount (int or list): инициализируемое или  прибавляемое значение"""

		if key in dictionary:
			dictionary[key] += amount
		else:
			dictionary[key] = amount

	@staticmethod
	def average(dictionary):
		"""Метод, считающий среднее целочиленное значение в переданном словаре
        Args:
            dictionary (dict): словарь с листом целочисленных значений
        Returns:
            dict (dict): словарь средних значений
        >>> DataSet('', '').average({2007: [10000, 15000, 20000], 2008: [30000, 35000, 40000]})
        {2007: 15000, 2008: 35000}
        """

		new_dictionary = {}
		for key, values in dictionary.items():
			new_dictionary[key] = int(sum(values) / len(values))
		return new_dictionary

	def csv_reader(self):
		"""Метод, читающий CSV файл и возвращающий итератор
        Returns:
            Iterator (dict): динамически сгенерированный словарь вакансии"""

		with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
			reader = csv.reader(file)
			header = next(reader)
			header_length = len(header)
			for row in reader:
				if '' not in row and len(row) == header_length:
					yield dict(zip(header, row))

	def get_statistic(self):
		"""Генерирует и возвращает статистику по вакансиям
        Returns:
            tuple[dict[int, int], dict[int, int], dict[int, int], dict[int, int],
            dict[string, int], dict[string, int]]:
            6 переменных, содержащих информацию о собранной статистике
        """

		salary = {}
		salary_of_vacancy_name = {}
		salary_city = {}
		count_of_vacancies = 0

		for vacancy_dictionary in self.csv_reader():
			vacancy = Vacancy(vacancy_dictionary)
			self.increment(salary, vacancy.year, [vacancy.salary_average])
			if vacancy.name.find(self.vacancy_name) != -1:
				self.increment(salary_of_vacancy_name, vacancy.year, [vacancy.salary_average])
			self.increment(salary_city, vacancy.area_name, [vacancy.salary_average])
			count_of_vacancies += 1

		vacancies_number = dict([(key, len(value)) for key, value in salary.items()])
		vacancies_number_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])

		if not salary_of_vacancy_name:
			salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
			vacancies_number_by_name = dict([(key, 0) for key, value in vacancies_number.items()])

		stats = self.average(salary)
		stats2 = self.average(salary_of_vacancy_name)
		stats3 = self.average(salary_city)

		stats4 = {}
		for year, salaries in salary_city.items():
			stats4[year] = round(len(salaries) / count_of_vacancies, 4)
		stats4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in stats4.items()]))
		stats4.sort(key=lambda a: a[-1], reverse=True)
		stats5 = stats4.copy()
		stats4 = dict(stats4)
		stats3 = list(filter(lambda a: a[0] in list(stats4.keys()), [(key, value) for key, value in stats3.items()]))
		stats3.sort(key=lambda a: a[-1], reverse=True)
		stats3 = dict(stats3[:10])
		stats5 = dict(stats5[:10])

		return stats, vacancies_number, stats2, vacancies_number_by_name, stats3, stats5

	@staticmethod
	def print_statistic(stats1, stats2, stats3, stats4, stats5, stats6):
		"""Метод, выводящий собранный данные
        Args:
            stats1 dict[int, int]: Динамика уровня зарплат по годам
            stats2 dict[int, int]: Динамика количества вакансий по годам
            stats3 dict[int, int]: Динамика уровня зарплат по годам для выбранной профессии
            stats4 dict[int, int]: Динамика количества вакансий по годам для выбранной профессии
            stats5 dict[str, int]: Уровень зарплат по городам (в порядке убывания)
            stats6 dict[str, int]: Доля вакансий по городам (в порядке убывания)"""

		print('Динамика уровня зарплат по годам: {0}'.format(stats1))
		print('Динамика количества вакансий по годам: {0}'.format(stats2))
		print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stats3))
		print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stats4))
		print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stats5))
		print('Доля вакансий по городам (в порядке убывания): {0}'.format(stats6))


if __name__ == '__main__':
	InCon()
	profile.disable()
	profile.print_stats(1)